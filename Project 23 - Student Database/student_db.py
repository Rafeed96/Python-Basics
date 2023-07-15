# Install packages pathlib
# install packge openpyxl
# install packge xlrd
# install packge pillow

from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

backG = "#06283D"
framBG = "#EDEDED"
framFG = "#06283D"

root=Tk()
root.title("Student Registration")
root.geometry("1250x7700+210+100")
root.configure(bg="#06283D")

file=pathlib.Path("Student_Data.xlsx")

if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet["A1"]="Registration Np."
    sheet["B1"]="Name"
    sheet["C1"]="Class"
    sheet["D1"]="Gender"
    sheet["E1"]="DOB"
    sheet["F1"]="Date of Registration"
    sheet["G1"]="Religion"
    sheet["H1"]="Skill"
    sheet["I1"]="Father's Name"
    sheet["J1"]="Mother's Name"
    sheet["K1"]="Father's Occupation"
    sheet["L1"]="Mother's Occupation"

    file.save("Student_Data.xlsx")

# Exit Button

def Exit():
    root.destroy()

# Show Image

def showImage():
    global filename
    global img
    filename=filedialog.askopenfilename(initialdir=os.getcwd(), title="Select Image File",filetypes=(("JPG File","*.jpg"),
                                                                                                     ("PNG File","*.png"),
                                                                                                     ("All Files","*.txt")))
    img = (Image.open(filename))
    resized_image = img.resize((190,190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2

# Registration Number

def registration_no():
    file=openpyxl.load_workbook("Student_Data.xlsx")
    sheet=file.active
    row=sheet.max_row

    max_row_value=sheet.cell(row=row,column=1).value

    try:
        Registration.set(max_row_value+1)
    except:
        Registration.set("1")




# Radio button for gender selection

def selection():
    value = radio.get()
    if value==1:
        gender = "Male"
    else:
        gender = "Female"


# Upper Frame
Label(root,text="Email: homepc@gmail.com",width=10,height=3,bg="#f0687c",anchor='e').pack(side=TOP,fill=X)
Label(root,text="STUDENT REGISTRATION",width=10,height=2,bg="#c36464",fg="#fff",font="arial 20 bold").pack(side=TOP,fill=X)


# Search box
Search=StringVar()
Entry(root,textvariable=Search,width=15,bd=2,font="arial 20").place(x=820,y=70)
imageicon3=PhotoImage(file="Project 23 - Student Database\images (1).png")
Srch=Button(root,text="Search",compound=LEFT,image=imageicon3,width=123,bg="#68ddfa",font="arial 13 bold")
Srch.place(x=1060,y=66)

imageicon4=PhotoImage(file="Project 23 - Student Database\imageupdate1.png")
Update_button=Button(root,image=imageicon4,bg="#c36464")
Update_button.place(x=110,y=64)


#Registration and Date
Label(root,text="Registration No.:", font="arial 13", fg=framBG,bg=backG,).place(x=30,y=150)
Label(root,text="Date:", font="arial 13", fg=framBG,bg=backG,).place(x=500,y=150)

Registration=StringVar()
Date=StringVar()

reg_entry = Entry(root,textvariable=Registration,width=15,font="arial 10")
reg_entry.place(x=160,y=150)

#Registration no.
registration_no()

today = date.today()
d1 = today.strftime("%m/%d/%Y")
date_entry = Entry(root,textvariable=Date,width=15,font="arial 10")
date_entry.place(x=550,y=150)

Date.set(d1)

#Student Details

obj=LabelFrame(root,text="Student's Details",font=20,bd=2,width=900,bg=framBG,height=250,relief=GROOVE)
obj.place(x=30,y=200)

#Column 1
Label(obj,text="Full Name:",font="arial 13", bg=framBG,fg=framFG).place(x=30,y=50)
Label(obj,text="Date Of Birth:",font="arial 13", bg=framBG,fg=framFG).place(x=30,y=100)
Label(obj,text="Gender:",font="arial 13", bg=framBG,fg=framFG).place(x=30,y=150)
#Column 2
Label(obj,text="Class:",font="arial 13", bg=framBG,fg=framFG).place(x=500,y=50)
Label(obj,text="Religion:",font="arial 13", bg=framBG,fg=framFG).place(x=500,y=100)
Label(obj,text="Skills:",font="arial 13", bg=framBG,fg=framFG).place(x=500,y=150)

Name=StringVar()
name_entry = Entry(obj,textvariable=Name,width=20,font="arial 10")
name_entry.place(x=160,y=50)

DateOB=StringVar()
dob_entry = Entry(obj,textvariable=DateOB,width=20,font="arial 10")
dob_entry.place(x=160,y=100)

radio = IntVar()
R1 = Radiobutton(obj,text="Male", variable=radio, value=1,bg=framBG,fg=framFG,command=selection)
R1.place(x=150,y=150)

R2 = Radiobutton(obj,text="Female", variable=radio, value=2,bg=framBG,fg=framFG,command=selection)
R2.place(x=200,y=150)

Religion=StringVar()
religion_entry = Entry(obj,textvariable=Religion,width=20,font="arial 10")
religion_entry .place(x=630,y=100)

Skill=StringVar()
skill_entry = Entry(obj,textvariable=Skill,width=20,font="arial 10")
skill_entry.place(x=630,y=150)

Student_Class = Combobox(obj,values=['1','2','3','4','5','6','7','8','9','10','11','12'],font="Roboto 10",width=17,state="r")
Student_Class.place(x=630,y=50)
Student_Class.set("Select Class")

#Parents Details

obj2=LabelFrame(root,text="Parent's Details",font=20,bd=2,width=900,bg=framBG,height=220,relief=GROOVE)
obj2.place(x=30,y=470)

Label(obj2,text="Father's Name:",font="arial 13", bg=framBG,fg=framFG).place(x=30,y=50)
Label(obj2,text="Occupation:",font="arial 13", bg=framBG,fg=framFG).place(x=30,y=100)

F_Name = StringVar()
f_entry=Entry(obj2,textvariable=F_Name,width=20,font="arial 10")
f_entry.place(x=160,y=50)

F_Occupation = StringVar()
fo_entry=Entry(obj2,textvariable=F_Occupation,width=20,font="arial 10")
fo_entry.place(x=160,y=100)

Label(obj2,text="Mother's Name:",font="arial 13", bg=framBG,fg=framFG).place(x=500,y=50)
Label(obj2,text="Occupation:",font="arial 13", bg=framBG,fg=framFG).place(x=500,y=100)

M_Name = StringVar()
m_entry=Entry(obj2,textvariable=M_Name,width=20,font="arial 10")
m_entry.place(x=630,y=50)

M_Occupation = StringVar()
mo_entry=Entry(obj2,textvariable=M_Occupation,width=20,font="arial 10")
mo_entry.place(x=630,y=100)


#Photo Image
f=Frame(root,bd=3,bg="black",width=200,height=200,relief=GROOVE)
f.place(x=1000,y=150)

img = PhotoImage(file="Project 23 - Student Database\image_profile.png")
lbl=Label(f,bg="black",image=img)
lbl.place(x=0,y=0)

# Buttons

Button(root,text="Upload",width=19,height=2,font="arial 12 bold",bg="lightblue",command=showImage).place(x=1000,y=370)

saveButton = Button(root,text="Save",width=19,height=2,font="arial 12 bold",bg="lightgreen").place(x=1000,y=450)

Button(root,text="Reset",width=19,height=2,font="arial 12 bold",bg="pink").place(x=1000,y=530)

Button(root,text="Exit",width=19,height=2,font="arial 12 bold",bg="gray",command=Exit).place(x=1000,y=610)




root.mainloop()
